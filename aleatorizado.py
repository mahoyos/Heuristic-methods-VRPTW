import os
import math
import random
import time
import openpyxl

def calculate_euclidean_distance(x1, y1, x2, y2):
    return round(math.sqrt((x1 - x2)**2 + (y1 - y2)**2), 3)

def process_file(file_path):
    with open(file_path, 'r') as file:
        n, Q = map(int, file.readline().strip().split())
        nodes = []
        for _ in range(n + 1):
            node_data = list(map(float, file.readline().strip().split()))
            nodes.append(node_data)
        
        nodes_dict = {}
        for i, node in enumerate(nodes):
            xi, yi, qi, ei, li, si = node[1:]
            nodes_dict[i] = {
                'x': xi,
                'y': yi,
                'demand': qi,
                'time_window': (ei, li),
                'service_time': si,
                'distances': {}
            }
        
        for i in range(n + 1):
            for j in range(n + 1):
                if i != j:
                    xi, yi = nodes[i][1], nodes[i][2]
                    xj, yj = nodes[j][1], nodes[j][2]
                    distance = calculate_euclidean_distance(xi, yi, xj, yj)
                    nodes_dict[i]['distances'][j] = distance
        
        return n, Q, nodes_dict


def find_insertion(route, unvisited, nodes_dict, capacity_left, current_time):
    feasible_nodes = []

    last_node = route[-1]

    for node in unvisited:
        if nodes_dict[node]['demand'] > capacity_left:
            continue

        arrival_time = current_time + nodes_dict[last_node]['distances'][node]
        if arrival_time > nodes_dict[node]['time_window'][1]:
            continue

        wait_time = max(nodes_dict[node]['time_window'][0] - arrival_time, 0)
        new_time = arrival_time + wait_time + nodes_dict[node]['service_time']
        
        return_to_depot_time = new_time + nodes_dict[node]['distances'][0]
        if return_to_depot_time > nodes_dict[0]['time_window'][1]:
            continue

        feasible_nodes.append((node, new_time))

    if not feasible_nodes:
        return None, None

    # Randomly select a feasible node
    selected_node, time = random.choice(feasible_nodes)

    return selected_node, time


def construct_routes(n, Q, nodes_dict):
    unvisited = set(range(1, n+1))
    routes = []
    total_distance = 0

    while unvisited:
        route = [0]
        capacity_left = Q
        current_time = 0
        route_distance = 0

        while True:
            node, time = find_insertion(route, unvisited, nodes_dict, capacity_left, current_time)
            if node is None:
                break

            last_node = route[-1]
            route_distance += nodes_dict[last_node]['distances'][node]
            route.append(node)
            unvisited.remove(node)
            capacity_left -= nodes_dict[node]['demand']
            current_time = time

            return_to_depot_time = current_time + nodes_dict[node]['distances'][0]
            if return_to_depot_time > nodes_dict[0]['time_window'][1]:
                route.pop()
                unvisited.add(node)
                break

        route_distance += nodes_dict[route[-1]]['distances'][0]
        total_distance += route_distance
        route.append(0)
        routes.append((route, route_distance))

    return routes, total_distance

def algorithm(n, Q, nodes_dict):
    start_time = time.time()

    routes, total_distance = construct_routes(n, Q, nodes_dict)
    computation_time = int((time.time() - start_time) * 1000)

    results = []
    results.append(f"{len(routes)} {round(total_distance, 3)} {computation_time}")

    for route, route_distance in routes:
        num_nodes = len(route) - 2  # Exclude the two depot visits
        current_time = 0
        load = 0
        times = []
        for i in range(len(route) - 1):
            node = route[i]
            next_node = route[i + 1]
            arrival_time = current_time + nodes_dict[node]['distances'][next_node]
            times.append(round(arrival_time, 3))
            if node != 0:  # If not the depot
                load += nodes_dict[node]['demand']
            current_time = arrival_time + nodes_dict[next_node]['service_time']

        results.append(f"{num_nodes} {' '.join(map(str, route))} {' '.join(map(str, times))} {int(load)}")

    return results

def save_to_excel(results_dict, filename):
    wb = openpyxl.Workbook()
    for instance_name, results in results_dict.items():
        ws = wb.create_sheet(title=instance_name)
        for i, line in enumerate(results, start=1):
            for j, value in enumerate(line.split(), start=1):
                ws.cell(row=i, column=j).value = value
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(filename)

def main():
    folder_path = ""  # Change this to the route with your files
    results_dict = {}
    for i in range(1, 19):
        file_name = f"VRPTW{i}.txt"
        instance_name = file_name.replace(".txt", "")
        file_path = os.path.join(folder_path, file_name)
        
        n, Q, nodes_dict = process_file(file_path)
        results = algorithm(n, Q, nodes_dict)
        results_dict[instance_name] = results

    save_to_excel(results_dict, "VRPTW_Miguel_Hoyos_aleatorizado.xlsx")

if __name__ == "__main__":
    main()
